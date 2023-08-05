from datetime import datetime

from openpyxl import Workbook, load_workbook

from openpyxl_templates.exceptions import OpenpyxlTemplateException
from openpyxl_templates.style import DefaultStyleSet, StyleSet
from openpyxl_templates.templated_sheet import TemplatedSheet
from openpyxl_templates.utils import OrderedType, Typed


class SheetnamesNotUnique(OpenpyxlTemplateException):
    def __init__(self, templated_workbook):
        super().__init__("Sheetnames are not unique on TemplatedWorkbook '%s'." % type(templated_workbook).__name__)


class MultipleActiveSheets(OpenpyxlTemplateException):
    def __init__(self, templated_workbook):
        super().__init__("The TemplatedWorkbook '%s' has multiple active sheets." % type(templated_workbook).__name__)


class TemplatedWorkbook(Workbook, metaclass=OrderedType):
    item_class = TemplatedSheet

    templated_sheets = None
    template_styles = Typed("template_styles", expected_type=StyleSet)

    timestamp = Typed("timestamp", expected_types=(str, bool), value=False)
    _default_timestamp = "%Y%m%d_%H%M%S"
    _file_extension = "xlsx"

    def __new__(cls, *args, file=None, **kwargs):
        if file:
            return load_workbook(file)
        return super().__new__(cls)

    def __init__(self, template_styles=None):
        super().__init__()

        self.template_styles = template_styles or self.template_styles or DefaultStyleSet()

        self.templated_sheets = []
        for attr, templated_sheet in self._items.items():
            if not templated_sheet._sheetname:
                templated_sheet._sheetname = attr

            templated_sheet.workbook = self
            self.templated_sheets.append(templated_sheet)

        self.validate()

    def validate(self):
        self._check_unique_sheetnames()
        self._check_only_one_active()

    def _check_unique_sheetnames(self):
        if len(set(templated_sheet.sheetname for templated_sheet in self.templated_sheets)) < len(self.templated_sheets):
            raise SheetnamesNotUnique(self)

    def _check_only_one_active(self):
        if len(tuple(sheet for sheet in self.templated_sheets if sheet.active)) > 1:
            raise MultipleActiveSheets(self)

    def remove_all_sheets(self):
        for sheetname in self.sheetnames:
            del self[sheetname]

    def save(self, filename):
        if self.timestamp:
            filename = self.timestamp_filename(filename)

        self.sort_worksheets()

        return super().save(filename)

    def sort_worksheets(self):
        order = {}
        index = 0
        active_index = 0
        for templated_sheet in self.templated_sheets:
            order[templated_sheet.sheetname] = index
            if templated_sheet.active:
               active_index = index
            index += 1

        for sheetname in self.sheetnames:
            if sheetname not in order:
                order[sheetname] = index
                index += 1

        self._sheets = sorted(self._sheets, key=lambda s: order[s.title])
        self.active = active_index

    def timestamp_filename(self, filename):
        return "%s_%s.%s" % (
            filename.strip(".%s" % self._file_extension),
            datetime.now().strftime(
                self.timestamp
                if isinstance(self.timestamp, str)
                else self._default_timestamp
            ),
            self._file_extension
        )


